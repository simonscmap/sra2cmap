#!/usr/bin/env python3
"""
Author : Ken Youens-Clark <kyclark@email.arizona.edu>
Date   : 2019-02-27
Purpose: Convert SRA metadata to CMAP import
"""

import argparse
import dateparser
import csv
import numpy as np
import os
import re
import sys
from collections import OrderedDict
from datetime import datetime
from openpyxl import Workbook


# --------------------------------------------------
def get_args():
    """get command-line arguments"""
    parser = argparse.ArgumentParser(
        description='Convert SRA metadata to CMAP import',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)

    parser.add_argument('sra', metavar='FILE', help='SRA metadata', nargs='+')

    parser.add_argument(
        '-d',
        '--delimiter',
        help='Field delimiter',
        metavar='str',
        type=str,
        default='\t')

    parser.add_argument(
        '-o',
        '--outdir',
        help='Output directory',
        metavar='str',
        type=str,
        default='export')

    return parser.parse_args()


# --------------------------------------------------
def warn(msg):
    """Print a message to STDERR"""
    print(msg, file=sys.stderr)


# --------------------------------------------------
def die(msg='Something bad happened'):
    """warn() and exit with error"""
    warn(msg)
    sys.exit(1)


# --------------------------------------------------
def format_record(rec):
    """Format the required fields: time, lat, lon, depth"""

    for fld in ['time', 'collection_date']:
        col_date = rec.get(fld)
        if col_date:
            dt = dateparser.parse(col_date)
            rec['time'] = dt.isoformat()

    # Combined e.g., 37.8305 S 41.1248 W
    # TODO: Handle other formats, e.g., HMS
    lat_lon = rec.get('lat_lon')
    if lat_lon:
        re1 = ('(\d+(?:\.\d+)?)'
               '\s*'
               '([NS])?'
               '(?:[,\s])'
               '(\d+(?:\.\d+)?)'
               '\s*'
               '([EW])?')
        match = re.search(re1, lat_lon)
        if match:
            lat, lat_dir, lon, lon_dir = match.groups()
            lat = float(lat)
            lon = float(lon)

            if lat_dir and lat_dir == 'S':
                lat *= -1

            if lon_dir and lon_dir == 'W':
                lon *= -1

            rec['lat'] = lat
            rec['lon'] = lon

    # TODO: verify that individual lat/lon fields are in proper format

    # Handles removing any non-numerical data, e.g., "9m"
    depth = rec.get('depth') or rec.get('sample_depth')
    if depth:
        match = re.search(r'(\d+(\.\d+)?)', depth)
        if match:
            rec['depth'] = match.group(1)

    return rec


# --------------------------------------------------
def normalize(s):
    """
    Make identifiers the same (snake_case_all_lower_case), e.g.:

    BioSample => bio_sample
    DATASTORE_filetype => datastore_filetype
    SRA_Study => sra_study
    """

    caps_under_re = re.compile(r'([A-Z]+)(?=[_])')
    s = caps_under_re.sub(lambda x: x.group(1).lower(), s)

    if '_' in s:
        s = '_'.join(map(str.lower, s.split('_')))
    else:
        camel_to_space = re.compile('((?<=[a-z0-9])[A-Z]|(?!^)[A-Z](?=[a-z]))')
        s = camel_to_space.sub(r'_\1', s).lower()

    return s


# --------------------------------------------------
def main():
    """Make a jazz noise here"""
    args = get_args()
    out_dir = args.outdir

    if out_dir and not os.path.isdir(out_dir):
        os.makedirs(out_dir)

    meta_flds = [
        'dataset_make', 'dataset_source', 'dataset_doi', 'dataset_history',
        'dataset_description', 'dataset_references'
    ]

    vars_flds = [
        'var_short_name', 'var_long_name', 'var_standard_name', 'var_unit',
        'var_sensor', 'var_spatial_res', 'var_temporal_res',
        'var_missing_value', 'var_discipline', 'var_keywords', 'var_comment'
    ]

    for i, file in enumerate(args.sra, start=1):
        if not os.path.isfile(file):
            warn('"{}" is not a file'.format(file))
            continue

        basename = os.path.basename(file)
        root, ext = os.path.splitext(basename)
        root = root.replace('_data', '')
        meta_file = os.path.join(os.path.dirname(file), root + '_meta' + ext)
        vars_file = os.path.join(os.path.dirname(file), root + '_vars' + ext)

        meta = {}
        if os.path.isfile(meta_file):
            for line in open(meta_file):
                field, val = line.rstrip('\n').split('\t')
                meta[field] = val

        vars_meta = {}
        if os.path.isfile(vars_file):
            for line in open(vars_file):
                field, val = line.rstrip('\n').split('\t')
                meta[field] = val

        print('{:3}: {:25}'.format(i, basename), end='')

        dir_name = out_dir if out_dir else os.path.dirname(file)
        out_file = os.path.join(dir_name, root + '.xlsx')
        wb = Workbook()

        req_flds = ['time', 'lat', 'lon', 'depth']
        with open(file) as fh:
            # "data" worksheet
            data_ws = wb.active
            data_ws.title = "data"
            reader = csv.DictReader(fh, delimiter=args.delimiter)
            flds = req_flds + reader.fieldnames
            norm2fld = OrderedDict(map(lambda s: (normalize(s), s), flds))
            ordered_flds = list(norm2fld.keys())
            data_ws.append(ordered_flds)

            num_taken = 0
            for row in reader:
                clean = format_record(row)
                if all([fld in row for fld in req_flds]):
                    num_taken += 1
                    data_ws.append(
                        [clean[norm2fld[fld]] for fld in ordered_flds])
                else:
                    print('Skipping\n', dict(clean))

            # "metadata" worksheet
            meta_ws = wb.create_sheet(title="dataset_meta_data")
            meta_ws.append(meta_flds)

            if meta:
                meta_ws.append(
                    list(map(lambda f: meta[f] if f in meta else '', meta_flds)))

            # "vars" worksheet
            vars_ws = wb.create_sheet(title="vars_meta_data")
            vars_ws.append(vars_flds)
            for fld in ordered_flds[4:]:
                pretty = ' '.join(fld.split('_')).title()
                unit = meta[fld] if fld in meta else ''
                vars_ws.append([fld, pretty, '', unit])

        wb.save(out_file)
        print('Exported {}'.format(num_taken))

    print('Done, see output in "{}".'.format(out_dir))


# --------------------------------------------------
if __name__ == '__main__':
    main()
